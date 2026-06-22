# -*- coding: utf-8 -*-
"""
excel_to_json.py
-----------------
Convierte apoyos.xlsx -> apoyos.json para el portal de Servicios y Programas.

CATEGORÍAS OFICIALES (las 13 que aparecen en la pestaña "Categorías" del
portal, con sus tarjetas e íconos propios). Un apoyo puede tener una o
varias, separadas por coma en la columna CATEGORÍA del Excel:

    Alimentación, Desarrollo Personal, Desarrollo Social, Discapacidad,
    Económico, Educación, Empleo, Mujeres, Niños - Niñas y Adolescentes,
    Personas en Vulnerabilidad, Personas Mayores, Transporte, Trámites

No existen las categorías "Salud" ni "General": un apoyo de tipo médico o
psicológico se clasifica como "Desarrollo Personal" (si es atención
individual) o "Desarrollo Social" (si es un servicio/actividad comunitaria).

QUÉ HACE EL SCRIPT:
  1. Lee cada fila de la hoja "APOYOS".
  2. Toma la(s) categoría(s) escritas a mano en la columna CATEGORÍA.
  3. Convierte el texto largo de REQUISITOS en una lista de viñetas limpia
     (detecta "a) b) c)", "1) 2) 3)", punto y coma, y secciones como
     "Requisitos generales:" / "Requisitos específicos:" / "Para el caso de...").

USO:
    1. Edita apoyos.xlsx (agrega, quita o modifica filas).
       - Si agregas un apoyo NUEVO, escribe su categoría (o varias, separadas
         por coma) en la columna CATEGORÍA usando los nombres de la lista
         de arriba exactamente. Si la dejas vacía, el apoyo no aparecerá en
         ningún filtro de categoría (solo en "Todos") hasta que la llenes.
    2. Corre:  python3 excel_to_json.py
    3. Reemplaza el apoyos.json en la carpeta del HTML con el nuevo.
    4. Recarga el navegador: el portal se actualiza solo, sin tocar código.

Requiere: pip install openpyxl
"""
import json
import re
import openpyxl

EXCEL_FILE = '../data/apoyos.xlsx'
JSON_FILE  = '../data/apoyos.json'
SHEET_NAME = 'APOYOS'

COLS = {
    'direccion_institucional': 'DIRECCIÓN INSTITUCIONAL',
    'nombre_apoyo':            'NOMBRE DEL APOYO/SERVICIO',
    'nombre_ciudadano':        'NOMBRE CIUDADANO',
    'categoria_raw':           'CATEGORÍA',
    'descripcion_oficial':     'DESCRIPCION DEL APOYO O SERVICIO',
    'descripcion_ciudadana':   'DESCRIPCIÓN CIUDADANA',
    'grupo_atencion':          'GRUPO DE ATENCIÓN',
    'municipio':               'MUNICIPIO',
    'genero':                  'GÉNERO',
    'etapa_vida':              'ETAPA DE VIDA',
    'especificaciones':        'ESPECIFICACIONES DEL APOYO',
    'monto':                   'MONTO DEL APOYO',
    'periodicidad':            'PERIODICIDAD DE ENTREGA',
    'donde_solicitar':         'DONDE SE SOLICITA',
    'requisitos':              'REQUISITOS',
    'donde_entrega':           'DONDE SE ENTREGA',
    'contacto':                'UBICACIÓN/CONTACTO',
    'observaciones':           'OBSERVACIONES',
}

# Etiqueta tal como se escribe en Excel (sin importar acentos/mayúsculas) -> clave interna
LABEL_TO_KEY = {
    'alimentacion': 'alimentacion', 'alimentación': 'alimentacion',
    'desarrollo personal': 'desarrollo_pers',
    'desarrollo social': 'desarrollo_soc',
    'discapacidad': 'discapacidad',
    'economico': 'economico', 'económico': 'economico',
    'educacion': 'capacitacion', 'educación': 'capacitacion',
    'capacitacion': 'capacitacion', 'capacitación': 'capacitacion',
    'empleo': 'empleo',
    'mujeres': 'mujeres',
    'niños, niñas y adolescentes': 'ninos', 'ninos, ninas y adolescentes': 'ninos',
    'niños': 'ninos', 'ninos': 'ninos', 'niñez': 'ninos',
    'personas en vulnerabilidad': 'vulnerabilidad', 'vulnerabilidad': 'vulnerabilidad',
    'personas mayores': 'mayores', 'mayores': 'mayores',
    'transporte': 'transporte',
    'tramites': 'tramites', 'trámites': 'tramites',
}

def clean(v):
    return '' if v is None else str(v).strip()

# Palabras clave (sinónimos y términos coloquiales) por categoría, para que
# la búsqueda encuentre apoyos aunque el ciudadano no use el nombre técnico
# (ej. "chamba" -> Empleo, "viejitos" -> Personas Mayores).
CATEGORY_KEYWORDS = {
    'alimentacion':    ['alimentación', 'comida', 'alimento', 'alimentos', 'despensa', 'despensas', 'canasta básica',
                         'nutrición', 'hambre', 'comer', 'víveres', 'comedor'],
    'desarrollo_pers': ['desarrollo personal', 'terapia', 'psicología', 'psicólogo', 'autoestima', 'bienestar emocional',
                         'consejería', 'ayuda psicológica', 'salud mental', 'depresión', 'ansiedad'],
    'desarrollo_soc':  ['desarrollo social', 'comunidad', 'convivencia', 'tejido social', 'vecinos', 'participación ciudadana',
                         'colonia', 'barrio'],
    'discapacidad':    ['discapacidad', 'silla de ruedas', 'invidente', 'sordo', 'down', 'autismo',
                         'capacidades diferentes', 'ortopédico', 'rehabilitación'],
    'economico':       ['económico', 'dinero', 'efectivo', 'apoyo económico', 'beca', 'subsidio', 'tarjeta',
                         'recurso económico', 'pago', 'ayuda económica'],
    'capacitacion':    ['educación', 'curso', 'cursos', 'taller', 'talleres', 'clases', 'aprender', 'oficio',
                         'capacitación', 'diploma', 'estudiar', 'beca educativa', 'escuela'],
    'empleo':          ['trabajo', 'chamba', 'vacante', 'vacantes', 'bolsa de trabajo', 'contratación',
                         'empleo', 'negocio', 'emprendimiento', 'proyecto productivo'],
    'mujeres':         ['mujer', 'mujeres', 'mamá', 'madre soltera', 'embarazo', 'embarazada',
                         'género', 'violencia de género'],
    'ninos':           ['niños', 'niñas', 'infancia', 'menores', 'adolescentes', 'jóvenes',
                         'juventud', 'escuela'],
    'vulnerabilidad':  ['vulnerabilidad', 'pobreza', 'necesidad', 'bajos recursos', 'escasos recursos', 'marginación',
                         'situación vulnerable', 'apoyo social'],
    'mayores':         ['personas mayores', 'ancianos', 'viejitos', 'tercera edad', 'jubilados', 'abuelos', 'abuelitos',
                         'adultos mayores', '60 años'],
    'transporte':      ['transporte', 'camión', 'autobús', 'traslado', 'pasaje', 'pasajes',
                         'movilidad', 'transporte público'],
    'tramites':        ['trámites', 'papeles', 'documentos', 'acta', 'actas', 'gratis', 'trámite',
                         'constancia', 'antecedentes penales', 'registro civil'],
}

def generate_keywords(categorias):
    seen = []
    for cat in categorias:
        for kw in CATEGORY_KEYWORDS.get(cat, []):
            if kw not in seen:
                seen.append(kw)
    return seen

def derive_direccion_corta(direccion):
    d = clean(direccion)
    dl = d.lower()
    if not d:
        return ''
    if 'cohesión social' in dl or 'cohesion social' in dl:
        return 'Cohesión Social'
    if 'grupos vulnerables' in dl:
        return 'Grupos Vulnerables'
    if 'frontera norte' in dl:
        return 'Subsecretaría Frontera Norte'
    m = re.search(r'"([^"]+)"', d)
    if m:
        return m.group(1).strip()
    return d

def normalize_municipio(texto):
    t = clean(texto)
    if not t:
        return t
    # unifica saltos de línea y espacios extra para que la lista de
    # municipios siempre quede separada por comas de forma consistente
    t = t.replace('\n', ', ')
    t = re.sub(r',\s*,', ',', t)
    t = re.sub(r'\s+', ' ', t)
    t = re.sub(r'\s*,\s*', ', ', t).strip(', ').strip()
    return t

# Nombres de municipios que contienen " y " como parte real del nombre
# (no deben separarse en dos al detectar listas "A, B y C").
MUNICIPIO_EXCEPTIONS = ['Guadalupe y Calvo']

def split_municipios(texto):
    """Convierte el texto de municipios en una lista limpia de nombres
    individuales, manejando tanto comas como el 'y' final de las listas
    en español ('A, B y C'), sin romper nombres compuestos como
    'Guadalupe y Calvo'."""
    t = clean(texto)
    if not t:
        return []
    tl = t.lower()
    if tl.startswith('todos los municipios') or tl.startswith('no especificado'):
        return [t]
    parts = [p.strip() for p in t.split(',') if p.strip()]
    if not parts:
        return [t]
    last = parts[-1]
    is_exception = any(last.strip().lower() == exc.lower() for exc in MUNICIPIO_EXCEPTIONS)
    if not is_exception and ' y ' in last:
        sub = [s.strip() for s in last.split(' y ', 1)]
        if len(sub) == 2 and all(sub):
            parts = parts[:-1] + sub
    return parts

def parse_categorias(raw):
    raw = clean(raw)
    if not raw:
        return []
    keys = []
    for part in raw.split(','):
        key = LABEL_TO_KEY.get(part.strip().lower())
        if key and key not in keys:
            keys.append(key)
    return keys

# ---------- parser de REQUISITOS en lista de viñetas ----------
def clean_item(s):
    s = s.strip()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r';\s*y\s*$', '', s, flags=re.IGNORECASE)
    s = s.rstrip(';,. ').strip()
    if not s:
        return ''
    if not s.endswith(('.', ':', '?', '!')):
        s += '.'
    return s[0].upper() + s[1:]

def split_markers(text):
    pattern = re.compile(r'(?:(?<=\n)|^)[ \t]*(?:[a-zA-Z]\)|\d+\))[ \t]*', re.MULTILINE)
    matches = list(pattern.finditer(text))
    if not matches:
        return None
    items = []
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip()
        if chunk:
            items.append(chunk)
    return items or None

def split_semicolons(text):
    parts = [p.strip() for p in re.split(r';\s*', text) if p.strip()]
    return parts if len(parts) >= 2 else None

SCENARIO_RE = re.compile(
    r'[Pp]ara el caso de(?:l)?\s+(?:la |el |los |las )?(.+?)'
    r'(?:,|:|\s+los\s+requisitos|\s+las\s+requisitos|\s+se\s+requiere|\s+es\s+necesario|'
    r'\s+deber[aá]n?|\s+son\s+los|\s+son\s+las)',
    re.IGNORECASE
)

def parse_scenarios(text):
    text = text.strip()
    if not text:
        return []
    matches = list(SCENARIO_RE.finditer(text))
    if matches:
        blocks = []
        lead = text[:matches[0].start()].strip()
        if lead and len(lead) > 8:
            litems = split_markers(lead) or split_semicolons(lead) or [lead]
            blocks.append({'titulo': None, 'items': [clean_item(i) for i in litems if i.strip()]})
        for i, m in enumerate(matches):
            title = re.sub(r'\s+', ' ', m.group(1).strip())
            title = title[0].upper() + title[1:]
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            chunk = text[start:end].strip()
            items = split_markers(chunk) or split_semicolons(chunk) or [chunk]
            items = [clean_item(it) for it in items if it.strip()]
            blocks.append({'titulo': title, 'items': items})
        return blocks

    items = split_markers(text)
    if items:
        return [{'titulo': None, 'items': [clean_item(it) for it in items]}]

    items = split_semicolons(text)
    if items:
        return [{'titulo': None, 'items': [clean_item(it) for it in items]}]

    return [{'titulo': None, 'items': [clean_item(text)]}]

SECTION_RE = re.compile(r'(Requisitos\s+general(?:es)?\s*:|Requisitos\s+espec[ií]fic[oa]s?\s*:)', re.IGNORECASE)

def parse_requisitos(text):
    text = clean(text)
    if not text:
        return []
    parts = SECTION_RE.split(text)
    if len(parts) > 1:
        blocks = []
        pre = parts[0].strip()
        if pre:
            blocks.extend(parse_scenarios(pre))
        i = 1
        while i < len(parts):
            header = parts[i].strip().rstrip(':')
            header = (header[0].upper() + header[1:].lower()) if header else header
            content = parts[i + 1] if i + 1 < len(parts) else ''
            i += 2
            for sb in parse_scenarios(content):
                title = f"{header} – {sb['titulo']}" if sb['titulo'] else header
                blocks.append({'titulo': title, 'items': sb['items']})
        return blocks
    return parse_scenarios(text)

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [clean(c.value) for c in ws[1]]
    col_idx = {}
    for field, header_name in COLS.items():
        if header_name not in headers:
            print(f"AVISO: no se encontró la columna '{header_name}'. Se omitirá ese campo.")
            continue
        col_idx[field] = headers.index(header_name)

    apoyos = []
    cats_usadas = set()
    sin_categoria = []
    next_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_idx = col_idx.get('nombre_apoyo')
        if nombre_idx is None or not clean(row[nombre_idx]):
            continue

        item = {'id': next_id}
        for field, idx in col_idx.items():
            if field == 'categoria_raw':
                continue
            item[field] = clean(row[idx])

        item['direccion_corta'] = derive_direccion_corta(item.get('direccion_institucional', ''))
        item['municipio'] = normalize_municipio(item.get('municipio', ''))
        item['municipios_lista'] = split_municipios(item['municipio'])

        cats = parse_categorias(row[col_idx['categoria_raw']]) if 'categoria_raw' in col_idx else []
        if not cats:
            sin_categoria.append(next_id)
        item['categorias'] = cats
        cats_usadas.update(cats)

        item['requisitos_estructurado'] = parse_requisitos(item.get('requisitos', ''))
        item['palabras_clave'] = generate_keywords(cats)

        apoyos.append(item)
        next_id += 1

    data = {
        'metadata': {
            'fuente': 'Secretaría de Desarrollo Humano y Bien Común — Chihuahua',
            'total': len(apoyos),
            'categorias': sorted(cats_usadas),
        },
        'apoyos': apoyos,
    }

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Listo: {len(apoyos)} apoyos exportados a {JSON_FILE}")
    print(f"Categorías detectadas: {sorted(cats_usadas)}")
    if sin_categoria:
        print(f"AVISO: estos ids no tienen categoría escrita en el Excel: {sin_categoria}")

if __name__ == '__main__':
    main()
